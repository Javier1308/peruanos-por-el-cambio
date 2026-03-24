import csv
import io
from datetime import timezone
from fastapi import APIRouter, Depends, HTTPException, Header, Query, Request
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, cast, Date
from app.database import get_db
from app.models.personero import Personero
from app.schemas.personero import PersoneroAdmin, StatsResponse
from app.config import settings

router = APIRouter(prefix="/admin", tags=["admin"])


def verify_admin(x_api_key: str = Header(...)):
    if x_api_key != settings.ADMIN_API_KEY:
        raise HTTPException(status_code=403, detail="Acceso no autorizado")
    return x_api_key


@router.get("/personeros", response_model=list[PersoneroAdmin])
async def listar_personeros(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    departamento: str | None = Query(None),
    dni_verificado: bool | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_admin),
):
    query = select(Personero).order_by(Personero.created_at.desc())

    if departamento:
        query = query.where(Personero.departamento == departamento)
    if dni_verificado is not None:
        query = query.where(Personero.dni_verificado == dni_verificado)

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    return result.scalars().all()


@router.get("/personeros/export/csv")
async def exportar_csv(
    departamento: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_admin),
):
    query = select(Personero).order_by(Personero.created_at.desc())
    if departamento:
        query = query.where(Personero.departamento == departamento)

    result = await db.execute(query)
    personeros = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Nombres", "Apellidos", "DNI", "Teléfono", "Email",
        "Departamento", "Provincia", "Distrito", "Local de Votación",
        "DNI Verificado", "IP Registro", "Fecha Registro"
    ])

    for p in personeros:
        writer.writerow([
            p.id, p.nombres, p.apellidos, p.dni, p.telefono, p.email,
            p.departamento, p.provincia, p.distrito,
            p.local_votacion or "",
            "Sí" if p.dni_verificado else "No",
            str(p.ip_registro),
            p.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=personeros.csv"},
    )


@router.get("/personeros/export/excel")
async def exportar_excel(
    departamento: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_admin),
):
    query = select(Personero).order_by(Personero.created_at.desc())
    if departamento:
        query = query.where(Personero.departamento == departamento)

    result = await db.execute(query)
    personeros = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personeros"

    headers = [
        "ID", "Código de Registro", "Nombres", "Apellidos", "DNI",
        "Teléfono", "Email", "Departamento", "Provincia", "Distrito",
        "Local de Votación", "DNI Verificado", "IP Registro", "Fecha Registro",
    ]

    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, p in enumerate(personeros, 2):
        created_at_naive = p.created_at.replace(tzinfo=None) if p.created_at.tzinfo else p.created_at
        ws.append([
            p.id,
            str(p.codigo_registro),
            p.nombres,
            p.apellidos,
            p.dni,
            p.telefono,
            p.email,
            p.departamento,
            p.provincia,
            p.distrito,
            p.local_votacion or "",
            "Sí" if p.dni_verificado else "No",
            str(p.ip_registro),
            created_at_naive,
        ])
        ws.cell(row=row, column=14).number_format = "YYYY-MM-DD HH:MM:SS"

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"personeros{'_' + departamento if departamento else ''}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/stats", response_model=StatsResponse)
async def estadisticas(
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_admin),
):
    total_result = await db.execute(select(func.count(Personero.id)))
    total = total_result.scalar_one()

    verificados_result = await db.execute(
        select(func.count(Personero.id)).where(Personero.dni_verificado == True)
    )
    verificados = verificados_result.scalar_one()

    dept_result = await db.execute(
        select(Personero.departamento, func.count(Personero.id))
        .group_by(Personero.departamento)
        .order_by(func.count(Personero.id).desc())
    )
    por_departamento = {row[0]: row[1] for row in dept_result.fetchall()}

    return StatsResponse(
        total=total,
        verificados=verificados,
        no_verificados=total - verificados,
        por_departamento=por_departamento,
    )
